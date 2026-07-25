"""KB-067: Generate PDF and Excel exports from customer-safe report sections."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _kv_lines(title: str, mapping: Dict[str, Any]) -> List[str]:
    lines = [title]
    for key, value in mapping.items():
        if isinstance(value, dict):
            lines.append(f"  {key}:")
            for k2, v2 in value.items():
                lines.append(f"    {k2}: {v2}")
        else:
            lines.append(f"  {key}: {value}")
    return lines


def build_pdf_bytes(
    *,
    title: str,
    executive_summary: str | None,
    published_at: str | None,
    sections: Dict[str, Any],
) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    body = styles["BodyText"]
    small = ParagraphStyle("Small", parent=body, fontSize=9, leading=12)

    story: List[Any] = []
    story.append(Paragraph("Kestrel Cyber — Monthly Security Report", h1))
    story.append(Paragraph(title, h2))
    cover = sections.get("cover") or {}
    period = sections.get("period") or {}
    story.append(
        Paragraph(
            f"Customer: {cover.get('customer_name', '—')} ({cover.get('short_code', '—')})<br/>"
            f"Period: {period.get('label') or period.get('report_month') or '—'}<br/>"
            f"SLA: {cover.get('sla_level', '—')} · Criticality: {cover.get('business_criticality', '—')}<br/>"
            f"Published: {published_at or '—'} · Snapshot: {sections.get('generated_at') or '—'}",
            small,
        )
    )
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("1. Executive Summary", h2))
    story.append(Paragraph((executive_summary or "No executive summary provided.").replace("\n", "<br/>"), body))
    story.append(Spacer(1, 0.15 * inch))

    posture = sections.get("posture") or {}
    story.append(Paragraph("2. Security Posture Snapshot", h2))
    posture_data = [
        ["Metric", "Value"],
        ["Appliances total", posture.get("appliances_total", 0)],
        ["Appliances online", posture.get("appliances_online", 0)],
        ["Appliances offline", posture.get("appliances_offline", 0)],
        ["Protected assets", posture.get("assets_total", 0)],
    ]
    for crit, n in (posture.get("assets_by_criticality") or {}).items():
        posture_data.append([f"Assets ({crit})", n])
    story.append(_table(posture_data))
    story.append(Spacer(1, 0.15 * inch))

    detection = sections.get("detection") or {}
    story.append(Paragraph("3. Detection Volume", h2))
    det_rows = [["Metric", "Value"], ["Alerts total", detection.get("alerts_total", 0)]]
    for k, v in (detection.get("by_severity") or {}).items():
        det_rows.append([f"Severity {k}", v])
    for k, v in (detection.get("by_status") or {}).items():
        det_rows.append([f"Status {k}", v])
    story.append(_table(det_rows))
    story.append(Spacer(1, 0.15 * inch))

    incidents = sections.get("incidents") or {}
    story.append(Paragraph("4. Incident Outcomes", h2))
    inc_rows = [
        ["Metric", "Value"],
        ["Opened this month", incidents.get("opened", 0)],
        ["Closed this month", incidents.get("closed", 0)],
        ["Still open (current)", incidents.get("still_open", 0)],
    ]
    for k, v in (incidents.get("by_severity_opened") or {}).items():
        inc_rows.append([f"Opened ({k})", v])
    story.append(_table(inc_rows))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph("Notable customer-visible incidents", styles["Heading3"]))
    notable = incidents.get("notable") or []
    if not notable:
        story.append(Paragraph("None recorded for this period.", body))
    else:
        for item in notable:
            story.append(
                Paragraph(
                    f"<b>{item.get('incident_number')}</b> [{item.get('severity')}/{item.get('status')}] "
                    f"{item.get('title')}<br/>{item.get('customer_visible_summary') or ''}",
                    small,
                )
            )
    story.append(Spacer(1, 0.15 * inch))

    recs = sections.get("recommendations") or {}
    story.append(Paragraph("5. Recommendations & Action Items", h2))
    story.append(
        Paragraph(
            f"Open (created in month): {recs.get('open_count', 0)} · "
            f"Completed (created in month): {recs.get('completed_count', 0)}",
            body,
        )
    )
    for item in recs.get("items") or []:
        story.append(
            Paragraph(
                f"• [{item.get('priority')}/{item.get('status')}] {item.get('title')} "
                f"({item.get('category')})",
                small,
            )
        )
    story.append(Spacer(1, 0.15 * inch))

    notif = sections.get("notifications") or {}
    story.append(Paragraph("6. Notification Activity", h2))
    story.append(
        Paragraph(
            f"Sent: {notif.get('sent_count', 0)} · Delivered: {notif.get('delivered_count', 0)}",
            body,
        )
    )
    story.append(Spacer(1, 0.15 * inch))

    narrative = sections.get("narrative") or {}
    story.append(Paragraph("7. Period Narrative", h2))
    story.append(Paragraph(f"<b>Highlights:</b> {narrative.get('period_highlights') or '—'}", body))
    story.append(Paragraph(f"<b>Trends:</b> {narrative.get('trends') or '—'}", body))
    story.append(Paragraph(f"<b>Next month focus:</b> {narrative.get('next_month_focus') or '—'}", body))
    story.append(Paragraph(f"<b>Leadership asks:</b> {narrative.get('leadership_asks') or '—'}", body))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(str(sections.get("deferred_kpis_note") or ""), small))

    doc.build(story)
    return buffer.getvalue()


def _table(data: List[List[Any]]) -> Table:
    table = Table(data, hAlign="LEFT", colWidths=[3.2 * inch, 2.2 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def build_xlsx_bytes(
    *,
    title: str,
    executive_summary: str | None,
    published_at: str | None,
    sections: Dict[str, Any],
) -> bytes:
    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    cover.append(["Field", "Value"])
    cover.append(["Title", title])
    cover.append(["Published", published_at or ""])
    cover.append(["Snapshot generated", sections.get("generated_at") or ""])
    for key, value in (sections.get("cover") or {}).items():
        cover.append([key, value])
    for key, value in (sections.get("period") or {}).items():
        cover.append([f"period.{key}", value])
    cover.append(["Executive summary", executive_summary or ""])

    posture = wb.create_sheet("Posture")
    posture.append(["Metric", "Value"])
    for key, value in (sections.get("posture") or {}).items():
        if isinstance(value, dict):
            for k2, v2 in value.items():
                posture.append([f"{key}.{k2}", v2])
        else:
            posture.append([key, value])

    detection = wb.create_sheet("Detection")
    detection.append(["Metric", "Value"])
    det = sections.get("detection") or {}
    detection.append(["alerts_total", det.get("alerts_total", 0)])
    for group in ("by_severity", "by_status"):
        for k, v in (det.get(group) or {}).items():
            detection.append([f"{group}.{k}", v])

    incidents = wb.create_sheet("Incidents")
    incidents.append(["Metric", "Value"])
    inc = sections.get("incidents") or {}
    for key in ("opened", "closed", "still_open"):
        incidents.append([key, inc.get(key, 0)])
    for k, v in (inc.get("by_severity_opened") or {}).items():
        incidents.append([f"opened.{k}", v])
    incidents.append([])
    incidents.append(
        ["incident_number", "title", "severity", "status", "customer_visible_summary"]
    )
    for item in inc.get("notable") or []:
        incidents.append(
            [
                item.get("incident_number"),
                item.get("title"),
                item.get("severity"),
                item.get("status"),
                item.get("customer_visible_summary"),
            ]
        )

    recs_sheet = wb.create_sheet("Recommendations")
    recs = sections.get("recommendations") or {}
    recs_sheet.append(["open_count", recs.get("open_count", 0)])
    recs_sheet.append(["completed_count", recs.get("completed_count", 0)])
    recs_sheet.append([])
    recs_sheet.append(["title", "priority", "status", "category", "due_at"])
    for item in recs.get("items") or []:
        recs_sheet.append(
            [
                item.get("title"),
                item.get("priority"),
                item.get("status"),
                item.get("category"),
                item.get("due_at"),
            ]
        )

    notif_sheet = wb.create_sheet("Notifications")
    notif = sections.get("notifications") or {}
    notif_sheet.append(["sent_count", notif.get("sent_count", 0)])
    notif_sheet.append(["delivered_count", notif.get("delivered_count", 0)])
    notif_sheet.append([])
    notif_sheet.append(["notification_type", "count"])
    for k, v in (notif.get("by_type") or {}).items():
        notif_sheet.append([k, v])

    narrative_sheet = wb.create_sheet("Narrative")
    narrative = sections.get("narrative") or {}
    for key in ("period_highlights", "trends", "next_month_focus", "leadership_asks"):
        narrative_sheet.append([key, narrative.get(key) or ""])
    narrative_sheet.append(["deferred_kpis_note", sections.get("deferred_kpis_note") or ""])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(short_code: str, report_month: str, ext: str) -> str:
    month = str(report_month)[:7].replace("-", "")
    return f"Kestrel_{short_code}_{month}_Monthly_Report.{ext}"
